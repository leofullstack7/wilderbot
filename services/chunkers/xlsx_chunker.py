from typing import List, Dict, Any
from io import BytesIO
from openpyxl import load_workbook
from utils.text_utils import normalize_text

def parse_xlsx(content_bytes: bytes, doc_title: str) -> List[Dict[str, Any]]:
    """
    Lee .xlsx solo con openpyxl.
    - Toma la primera fila como encabezados.
    - Cada fila siguiente => 1 item "Col: valor; Col2: valor2..."
    - Máx 15 columnas por simplicidad.
    """
    items: List[Dict[str, Any]] = []
    wb = load_workbook(BytesIO(content_bytes), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue

            # Encabezados (primera fila con algo de contenido)
            headers = None
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                raw = list(row[:15])
                if any(c is not None and str(c).strip() != "" for c in raw):
                    headers = [
                        (str(c).strip() if c is not None and str(c).strip() != "" else f"Col{idx+1}")
                        for idx, c in enumerate(raw)
                    ]
                break
            if not headers:
                continue

            # Filas de datos
            for r in ws.iter_rows(min_row=2, values_only=True):
                cells = list(r[:15])
                if all(c is None or str(c).strip() == "" for c in cells):
                    continue

                pairs = []
                for h, v in zip(headers, cells):
                    if v is None:
                        continue
                    txt = normalize_text(str(v))
                    if txt:
                        pairs.append(f"{h}: {txt}")

                row_text = "; ".join(pairs)
                row_text = normalize_text(row_text)
                if row_text:
                    items.append({
                        "text": row_text,
                        "metadata": {
                            "doc_title": doc_title,
                            "source_type": "excel",
                            "sheet": ws.title
                        }
                    })
    finally:
        wb.close()
    return items
